"""
Transform an extracted Air Rate Card dataframe into a matrix-format workbook.

Reads from the processing folder and writes the matrix rate card to the output folder.
"""

from __future__ import annotations

import re
from collections.abc import Callable
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from project_paths import OUTPUT_DIR, PROCESSING_DIR

COST_NAME_ROW = 1
APPLY_IF_ROW = 2
RATE_BY_ROW = 3
WEIGHT_BRACKET_ROW = 4
COLUMN_HEADER_ROW = 5
DATA_START_ROW = 6

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(bold=True, color="1F4E78")
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4B4B4"),
    right=Side(style="thin", color="B4B4B4"),
    top=Side(style="thin", color="B4B4B4"),
    bottom=Side(style="thin", color="B4B4B4"),
)
PRICE_NUMBER_FORMAT = "#,##0.00"
DATE_NUMBER_FORMAT = "DD/MM/YYYY"

DATE_SHIPMENT_COLUMNS = {
    "ValidFrom",
    "ValidUntil",
}

HEADER_ROW_HEIGHT = 18
DATA_ROW_HEIGHT = 15

VOLUME_WEIGHT_RATIO_COLUMN = "volume weight ratio"
REMARKS_COLUMN = "Remarks"
DROP_SHIPMENT_REMARK_MARKER = "drop-shipment"
FIX_FUEL_SURCHARGE_COLUMN = "fix fuel surcharge"
FIX_SECURITY_SURCHARGE_COLUMN = "fix security surcharge"
SECURITY_SURCH_OUTLAY_COLUMN = "security surch outlay"
FUEL_SURCHARGE_OUTLAY_COLUMN = "fuel surcharge outlay"
FUEL_SURCHARGE_COST_NAME = "Fuel Surcharge"
SECURITY_SURCHARGE_COST_NAME = "Security Surcharge"
DGR_FEE_COLUMN = "DGR fee"
DGR_FEE_HEADER = "DGR Fee"
DGR_FEE_SHIPMENT_HEADER = 'DGR fee "yes" = as per outlay amount = fix'
FLAT_IPC_COLUMN = "Flat IPC"
FLAT_TRANSPORT_COLUMN = "Flat"
ORIGIN_COUNTRY_CODE_COLUMN = "origin country code"
DESTINATION_COUNTRY_CODE_COLUMN = "destination country code"
LATAM_COUNTRY_CODES = ("BR", "CL", "CO", "EC", "MX", "PE")

BOLD_SHIPMENT_HEADERS = {
    "Origin Zone",
    "Destination Zone",
    "service code extended",
    "Deck",
    "Airline",
}

SHIPMENT_PLACEHOLDER_AFTER = {
    "Airdeck": "Deck",
    "airline": "Airline",
}

VOLUME_WEIGHT_RATIO_LABELS = {
    0: "0:0",
    6: "1:6",
    7: "1:7",
    8: "1:8",
}

CURRENCY_HELPER_COLUMNS = {
    "Currency",
    "Currency Outbound",
    "Currency Inbound",
}

CURRENCY_BY_COST_PREFIX = {
    "Flat OPC": "Currency Outbound",
    "Flat IPC": "Currency Inbound",
}

EXCLUDED_SHIPMENT_COLUMNS = {
    "country contract owner",
    "published P PC LO int",
    "forwarder name",
    "origin country",
    "origin city",
    "destination country",
    "destination city",
    "transit time in hours",
    "max length cm",
    "max width cm",
    "max height cm",
    "frequency",
    "date of update",
    "responsible person",
    "origin apt zone cluster",
    "dest apt zone cluster",
    "Rate scope",
    "Status",
    "Version",
}

RATIO_HANDLED_COST_COLUMNS = {
    FIX_FUEL_SURCHARGE_COLUMN,
    FIX_SECURITY_SURCHARGE_COLUMN,
    SECURITY_SURCH_OUTLAY_COLUMN,
}


@dataclass
class ShipmentColumn:
    header: str
    source_column: str | None = None
    bold_header: bool = False


@dataclass
class MatrixCostColumn:
    cost_name: str
    source_column: str
    bracket_label: str
    rate_unit: str
    is_currency: bool = False
    header_label: str | None = None
    parse_value: Callable[[object], object] | None = None


@dataclass
class MatrixCostBlock:
    cost_name: str
    currency_column: str
    columns: list[MatrixCostColumn] = field(default_factory=list)
    volume_weight_ratio: object | None = None
    drop_shipment_only: bool = False


@dataclass
class MatrixBuildResult:
    matrix_path: Path
    row_count: int
    shipment_column_count: int
    cost_block_count: int


def _normalize_label(label: str) -> str:
    return re.sub(r"\s+", " ", str(label).replace("\n", " ")).strip()


def _column_key(label: str) -> str:
    return _normalize_label(label).lower()


def is_transport_cost_column(column: str) -> bool:
    lower = _column_key(column)
    if any(token in lower for token in ("fee", "surcharge", "surch")):
        return False
    if "flat" in lower and lower != "flat":
        return False
    if lower in {"minimum", "flat"}:
        return True
    if re.search(r"\d+\s*kg", lower):
        return True
    return False


def is_cost_column(column: str) -> bool:
    lower = _column_key(column)
    if lower in {_column_key(name) for name in CURRENCY_HELPER_COLUMNS}:
        return False
    if any(token in lower for token in ("fee", "surcharge", "surch")):
        return True
    if lower in {"minimum", "flat"}:
        return True
    if re.search(r"\d+\s*kg", lower):
        return True
    if "flat" in lower and lower != "flat":
        return True
    return False


def _parse_weight_bracket(column: str) -> str:
    lower = _column_key(column)
    if lower == "minimum":
        return "MIN"
    match = re.search(r"(\d+)\s*kg", lower)
    if match:
        return f"<={match.group(1)}"
    return ""


def _get_rate_unit(column: str) -> str:
    lower = _column_key(column)
    if re.search(r"\d+\s*kg", lower):
        return "p/unit"
    return "Flat"


def _format_volume_weight_ratio(value: object) -> str:
    numeric = int(float(value))
    if numeric in VOLUME_WEIGHT_RATIO_LABELS:
        return VOLUME_WEIGHT_RATIO_LABELS[numeric]
    if numeric == 0:
        return "0:0"
    return f"1:{numeric}"


def _ratio_cost_suffix(ratio_value: object) -> str:
    label = _format_volume_weight_ratio(ratio_value)
    return f"(volume/weight ratio {label})"


def _drop_shipment_cost_suffix(ratio_value: object) -> str:
    label = _format_volume_weight_ratio(ratio_value)
    return f"(drop-shipments, volume/weight ratio {label})"


def _is_drop_shipment_row(row: pd.Series) -> bool:
    if REMARKS_COLUMN not in row.index:
        return False
    remark = row[REMARKS_COLUMN]
    if _is_empty_value(remark):
        return False
    return DROP_SHIPMENT_REMARK_MARKER in str(remark).lower()


def _row_matches_block(row: pd.Series, block: MatrixCostBlock) -> bool:
    is_drop_shipment = _is_drop_shipment_row(row)
    if block.drop_shipment_only:
        if not is_drop_shipment:
            return False
    elif is_drop_shipment:
        return False

    if block.volume_weight_ratio is not None:
        return _row_matches_volume_weight_ratio(row, block.volume_weight_ratio)
    return True


def _column_has_numeric_values(rate_card: pd.DataFrame, column: str) -> bool:
    if column not in rate_card.columns:
        return False
    return pd.to_numeric(rate_card[column], errors="coerce").notna().any()


def _should_include_cost_column(column: str, rate_card: pd.DataFrame) -> bool:
    name = _normalize_label(column)
    if _column_key(name) == _column_key(FUEL_SURCHARGE_OUTLAY_COLUMN):
        return _column_has_numeric_values(rate_card, name)
    return is_cost_column(name)


def _currency_column_for_block(block: MatrixCostBlock) -> str:
    for prefix, currency_column in CURRENCY_BY_COST_PREFIX.items():
        if block.cost_name.startswith(prefix):
            return currency_column
    return "Currency"


def _is_date_shipment_column(header: str) -> bool:
    return header in DATE_SHIPMENT_COLUMNS


def _format_shipment_cell_value(header: str, value: object) -> object:
    if not _is_date_shipment_column(header):
        return value
    if _is_empty_value(value):
        return None

    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _build_shipment_columns(columns: pd.Index, rate_card: pd.DataFrame) -> list[ShipmentColumn]:
    shipment_columns: list[ShipmentColumn] = []
    for column in columns:
        name = _normalize_label(column)
        if _column_key(name) in {_column_key(item) for item in EXCLUDED_SHIPMENT_COLUMNS}:
            continue
        if is_cost_column(name):
            continue
        if name in CURRENCY_HELPER_COLUMNS:
            continue

        shipment_columns.append(
            ShipmentColumn(
                header=name,
                source_column=name,
                bold_header=name in BOLD_SHIPMENT_HEADERS,
            )
        )

        placeholder_header = SHIPMENT_PLACEHOLDER_AFTER.get(name)
        if placeholder_header is not None:
            shipment_columns.append(
                ShipmentColumn(
                    header=placeholder_header,
                    source_column=None,
                    bold_header=True,
                )
            )

    return shipment_columns


def _insert_dgr_fee_shipment_column(
    shipment_columns: list[ShipmentColumn],
    rate_card: pd.DataFrame,
) -> list[ShipmentColumn]:
    if DGR_FEE_COLUMN not in rate_card.columns:
        return shipment_columns

    updated = list(shipment_columns)
    insert_at = next(
        (
            index
            for index, column in enumerate(updated)
            if _column_key(column.header) == _column_key("Airdeck")
        ),
        len(updated),
    )
    updated.insert(
        insert_at,
        ShipmentColumn(
            header=DGR_FEE_SHIPMENT_HEADER,
            source_column=DGR_FEE_COLUMN,
            bold_header=False,
        ),
    )
    return updated


def _build_matrix_cost_column(
    cost_name: str,
    source_column: str,
    rate_unit: str | None = None,
) -> MatrixCostColumn:
    return MatrixCostColumn(
        cost_name=cost_name,
        source_column=source_column,
        bracket_label=_parse_weight_bracket(source_column),
        rate_unit=rate_unit or _get_rate_unit(source_column),
    )


def _transport_source_columns(columns: pd.Index) -> list[str]:
    return [
        _normalize_label(column)
        for column in columns
        if is_transport_cost_column(_normalize_label(column))
    ]


def _get_volume_weight_ratios(
    rate_card: pd.DataFrame,
    *,
    drop_shipment_only: bool = False,
) -> list[object]:
    if VOLUME_WEIGHT_RATIO_COLUMN not in rate_card.columns:
        return []

    rows = rate_card
    if drop_shipment_only:
        rows = rate_card[rate_card.apply(_is_drop_shipment_row, axis=1)]

    values = rows[VOLUME_WEIGHT_RATIO_COLUMN].dropna().unique().tolist()
    return sorted(values, key=lambda value: int(float(value)))


def _row_matches_volume_weight_ratio(
    row: pd.Series,
    ratio_value: object | None,
) -> bool:
    if ratio_value is None:
        return True
    if VOLUME_WEIGHT_RATIO_COLUMN not in row.index:
        return False

    row_value = row[VOLUME_WEIGHT_RATIO_COLUMN]
    if pd.isna(row_value):
        return False

    return int(float(row_value)) == int(float(ratio_value))


def _filter_rows_for_ratio(
    rate_card: pd.DataFrame,
    ratio_value: object,
    *,
    drop_shipment_only: bool = False,
    exclude_drop_shipment: bool = False,
) -> pd.DataFrame:
    ratio_mask = rate_card[VOLUME_WEIGHT_RATIO_COLUMN].apply(
        lambda value: not pd.isna(value)
        and int(float(value)) == int(float(ratio_value))
    )
    rows = rate_card[ratio_mask]

    if drop_shipment_only:
        return rows[rows.apply(_is_drop_shipment_row, axis=1)]
    if exclude_drop_shipment:
        return rows[~rows.apply(_is_drop_shipment_row, axis=1)]
    return rows


def _column_has_data_in_rows(column: str, rows: pd.DataFrame) -> bool:
    if column not in rows.columns:
        return False
    return rows[column].apply(lambda value: not _is_empty_value(value)).any()


def _is_weight_bracket_transport_column(column: str) -> bool:
    return bool(re.search(r"\d+\s*kg", _column_key(column)))


def _get_canonical_transport_columns(
    columns: pd.Index,
    rate_card: pd.DataFrame,
) -> list[str]:
    transport_columns = _transport_source_columns(columns)
    # Keep configured transport brackets even when source cells are empty;
    # _resolve_transport_column_value() will backfill bracket values from Flat.
    canonical = [
        column
        for column in transport_columns
        if _column_key(column) != _column_key(FLAT_TRANSPORT_COLUMN)
    ]
    if canonical:
        return canonical

    # Fallback for unusual layouts with only Flat or sparse columns.
    non_drop_rows = rate_card[~rate_card.apply(_is_drop_shipment_row, axis=1)]
    return [
        column
        for column in transport_columns
        if _column_has_data_in_rows(column, non_drop_rows)
    ]


def _resolve_transport_column_value(row: pd.Series, source_column: str) -> object:
    if source_column not in row.index:
        return None

    value = row[source_column]
    if not _is_empty_value(value):
        return value

    if (
        _is_weight_bracket_transport_column(source_column)
        and FLAT_TRANSPORT_COLUMN in row.index
    ):
        flat_value = row[FLAT_TRANSPORT_COLUMN]
        if not _is_empty_value(flat_value):
            return flat_value

    return None


def _is_transport_cost_block(block: MatrixCostBlock) -> bool:
    return block.cost_name.startswith("Transport cost")


def _build_transport_cost_block(
    cost_name: str,
    transport_columns: list[str],
    ratio_value: object,
    *,
    drop_shipment_only: bool,
) -> MatrixCostBlock | None:
    if not transport_columns:
        return None

    return MatrixCostBlock(
        cost_name=cost_name,
        currency_column="Currency",
        columns=[
            _build_matrix_cost_column(cost_name, column)
            for column in transport_columns
        ],
        volume_weight_ratio=ratio_value,
        drop_shipment_only=drop_shipment_only,
    )


def _build_fuel_and_security_blocks(
    rate_card: pd.DataFrame,
    ratio_value: object,
    suffix: str,
) -> list[MatrixCostBlock]:
    blocks: list[MatrixCostBlock] = []

    if FIX_FUEL_SURCHARGE_COLUMN in rate_card.columns:
        blocks.append(
            MatrixCostBlock(
                cost_name=f"{FUEL_SURCHARGE_COST_NAME} {suffix}",
                currency_column="Currency",
                columns=[
                    _build_matrix_cost_column(
                        f"{FUEL_SURCHARGE_COST_NAME} {suffix}",
                        FIX_FUEL_SURCHARGE_COLUMN,
                        rate_unit="p/unit",
                    )
                ],
                volume_weight_ratio=ratio_value,
                drop_shipment_only=False,
            )
        )

    security_source_column = None
    if FIX_SECURITY_SURCHARGE_COLUMN in rate_card.columns:
        security_source_column = FIX_SECURITY_SURCHARGE_COLUMN
    elif SECURITY_SURCH_OUTLAY_COLUMN in rate_card.columns:
        security_source_column = SECURITY_SURCH_OUTLAY_COLUMN

    if security_source_column is not None:
        blocks.append(
            MatrixCostBlock(
                cost_name=f"{SECURITY_SURCHARGE_COST_NAME} {suffix}",
                currency_column="Currency",
                columns=[
                    _build_matrix_cost_column(
                        f"{SECURITY_SURCHARGE_COST_NAME} {suffix}",
                        security_source_column,
                        rate_unit="p/unit",
                    )
                ],
                volume_weight_ratio=ratio_value,
                drop_shipment_only=False,
            )
        )

    return blocks


def _ratio_has_matching_rows(
    rate_card: pd.DataFrame,
    ratio_value: object,
    *,
    drop_shipment_only: bool,
    exclude_drop_shipment: bool,
) -> bool:
    return not _filter_rows_for_ratio(
        rate_card,
        ratio_value,
        drop_shipment_only=drop_shipment_only,
        exclude_drop_shipment=exclude_drop_shipment,
    ).empty


def _build_ratio_cost_blocks(
    columns: pd.Index,
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    blocks: list[MatrixCostBlock] = []
    canonical_transport_columns = _get_canonical_transport_columns(columns, rate_card)
    ratio_values = _get_volume_weight_ratios(rate_card)

    for ratio_value in ratio_values:
        suffix = _ratio_cost_suffix(ratio_value)

        if _ratio_has_matching_rows(
            rate_card,
            ratio_value,
            drop_shipment_only=False,
            exclude_drop_shipment=True,
        ):
            standard_transport = _build_transport_cost_block(
                f"Transport cost {suffix}",
                canonical_transport_columns,
                ratio_value,
                drop_shipment_only=False,
            )
            if standard_transport is not None:
                blocks.append(standard_transport)

        if _ratio_has_matching_rows(
            rate_card,
            ratio_value,
            drop_shipment_only=True,
            exclude_drop_shipment=False,
        ):
            drop_transport = _build_transport_cost_block(
                f"Transport cost {_drop_shipment_cost_suffix(ratio_value)}",
                canonical_transport_columns,
                ratio_value,
                drop_shipment_only=True,
            )
            if drop_transport is not None:
                blocks.append(drop_transport)

        if _ratio_has_matching_rows(
            rate_card,
            ratio_value,
            drop_shipment_only=False,
            exclude_drop_shipment=True,
        ):
            blocks.extend(_build_fuel_and_security_blocks(rate_card, ratio_value, suffix))

    return blocks


def _build_unified_cost_blocks(
    columns: pd.Index,
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    blocks: list[MatrixCostBlock] = []
    canonical_transport_columns = _get_canonical_transport_columns(columns, rate_card)
    transport_block = _build_transport_cost_block(
        "Transport cost",
        canonical_transport_columns,
        ratio_value=None,
        drop_shipment_only=False,
    )
    if transport_block is not None:
        blocks.append(transport_block)

    if FIX_FUEL_SURCHARGE_COLUMN in rate_card.columns:
        blocks.append(
            MatrixCostBlock(
                cost_name=FUEL_SURCHARGE_COST_NAME,
                currency_column="Currency",
                columns=[
                    _build_matrix_cost_column(
                        FUEL_SURCHARGE_COST_NAME,
                        FIX_FUEL_SURCHARGE_COLUMN,
                        rate_unit="p/unit",
                    )
                ],
            )
        )

    if SECURITY_SURCH_OUTLAY_COLUMN in rate_card.columns:
        blocks.append(
            MatrixCostBlock(
                cost_name=SECURITY_SURCHARGE_COST_NAME,
                currency_column="Currency",
                columns=[
                    _build_matrix_cost_column(
                        SECURITY_SURCHARGE_COST_NAME,
                        SECURITY_SURCH_OUTLAY_COLUMN,
                        rate_unit="p/unit",
                    )
                ],
            )
        )
    return blocks


def _is_ratio_handled_column(column: str) -> bool:
    name = _normalize_label(column)
    if is_transport_cost_column(name):
        return True
    return name in RATIO_HANDLED_COST_COLUMNS


def _build_standalone_cost_blocks(
    columns: pd.Index,
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    blocks: list[MatrixCostBlock] = []

    for column in columns:
        name = _normalize_label(column)
        if not _should_include_cost_column(name, rate_card):
            continue
        if _is_ratio_handled_column(name):
            continue
        if _is_dgr_fee_column(name):
            blocks.append(_build_dgr_fee_cost_block())
            continue
        if _is_flat_ipc_column(name):
            blocks.append(_build_flat_ipc_cost_block())
            continue

        blocks.append(
            MatrixCostBlock(
                cost_name=name,
                currency_column=_currency_column_for_block(
                    MatrixCostBlock(cost_name=name, currency_column="Currency")
                ),
                columns=[_build_matrix_cost_column(name, name)],
            )
        )

    return blocks


def _build_cost_blocks(
    columns: pd.Index,
    rate_card: pd.DataFrame,
    *,
    ignore_volume_weight_ratio: bool = False,
) -> list[MatrixCostBlock]:
    ratio_or_unified_blocks = (
        _build_unified_cost_blocks(columns, rate_card)
        if ignore_volume_weight_ratio
        else _build_ratio_cost_blocks(columns, rate_card)
    )
    return ratio_or_unified_blocks + _build_standalone_cost_blocks(
        columns,
        rate_card,
    )


def _is_empty_value(value: object) -> bool:
    if value is None or pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _sanitize_cost_value(value: object, *, allow_string: bool = False) -> object:
    if _is_empty_value(value):
        return None
    if allow_string:
        return value
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value == int(value):
            return int(value)
        return float(value)

    numeric = pd.to_numeric(str(value).replace(",", "."), errors="coerce")
    if pd.isna(numeric):
        return None
    if numeric == int(numeric):
        return int(numeric)
    return float(numeric)


def _is_zero_like(value: object) -> bool:
    if value is None or pd.isna(value):
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return float(value) == 0.0
    numeric = pd.to_numeric(str(value).strip().replace(",", "."), errors="coerce")
    return not pd.isna(numeric) and float(numeric) == 0.0


def _extract_numeric_token(text: str) -> int | float | None:
    cleaned = re.sub(
        r"^(?:EUR|USD|GBP|CHF|\$|€)\s*",
        "",
        text.strip(),
        flags=re.IGNORECASE,
    ).strip()
    match = re.search(r"(\d+(?:[.,]\d+)?)", cleaned)
    if not match:
        return None

    numeric = float(match.group(1).replace(",", "."))
    if numeric == int(numeric):
        return int(numeric)
    return numeric


def _parse_dgr_fee_parts(value: object) -> tuple[object, object]:
    if _is_empty_value(value):
        return None, None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if numeric == int(numeric):
            return int(numeric), None
        return numeric, None

    text = str(value).strip()
    if "plus" not in text.lower():
        numeric = _extract_numeric_token(text)
        return numeric, None

    first_part, second_part = re.split(r"\s+plus\s+", text, maxsplit=1, flags=re.IGNORECASE)
    first_value = _extract_numeric_token(first_part)
    second_value = _extract_numeric_token(second_part.split("/")[0])
    return first_value, second_value


def _parse_dgr_fee_primary(value: object) -> object:
    return _parse_dgr_fee_parts(value)[0]


def _parse_dgr_fee_secondary(value: object) -> object:
    return _parse_dgr_fee_parts(value)[1]


def _is_dgr_fee_column(column: str) -> bool:
    return _column_key(column) == _column_key(DGR_FEE_COLUMN)


def _is_flat_ipc_column(column: str) -> bool:
    return _column_key(column) == _column_key(FLAT_IPC_COLUMN)


def _is_dgr_fee_block(block: MatrixCostBlock) -> bool:
    return block.cost_name == DGR_FEE_HEADER


def _build_dgr_fee_cost_block() -> MatrixCostBlock:
    return MatrixCostBlock(
        cost_name=DGR_FEE_HEADER,
        currency_column="Currency",
        columns=[
            MatrixCostColumn(
                cost_name=DGR_FEE_HEADER,
                source_column="Currency",
                bracket_label="Currency",
                rate_unit="Currency",
                is_currency=True,
            ),
            MatrixCostColumn(
                cost_name=DGR_FEE_HEADER,
                source_column=DGR_FEE_COLUMN,
                bracket_label="",
                rate_unit="Flat",
                parse_value=_parse_dgr_fee_primary,
            ),
            MatrixCostColumn(
                cost_name=DGR_FEE_HEADER,
                source_column="Currency",
                bracket_label="Currency",
                rate_unit="Currency",
                is_currency=True,
            ),
            MatrixCostColumn(
                cost_name=DGR_FEE_HEADER,
                source_column=DGR_FEE_COLUMN,
                bracket_label="",
                rate_unit="p/unit",
                parse_value=_parse_dgr_fee_secondary,
            ),
        ],
    )


def _parse_flat_ipc_parts(value: object) -> tuple[object, object]:
    if _is_empty_value(value):
        return None, None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if numeric == int(numeric):
            return int(numeric), None
        return numeric, None

    text = str(value).strip()
    if not text:
        return None, None

    per_unit_value: int | float | None = None
    max_flat_value: int | float | None = None

    per_kg_match = re.search(r"(\d+(?:[.,]\d+)?)\s*per\s*kg", text, flags=re.IGNORECASE)
    if per_kg_match:
        per_unit_value = _extract_numeric_token(per_kg_match.group(1))

    max_match = re.search(
        r"maximum(?:\s+of)?\s+(?:EUR|USD|GBP|CHF|\$|€)?\s*(\d+(?:[.,]\d+)?)",
        text,
        flags=re.IGNORECASE,
    )
    if max_match:
        max_flat_value = _extract_numeric_token(max_match.group(1))

    if per_unit_value is None and max_flat_value is None:
        # Default IPC values should populate p/unit when no explicit maximum pattern exists.
        per_unit_value = _extract_numeric_token(text)
    elif per_unit_value is None and max_flat_value is not None:
        # If only a maximum pattern is present, keep it as MAX Flat.
        pass

    return per_unit_value, max_flat_value


def _parse_flat_ipc_per_unit(value: object) -> object:
    return _parse_flat_ipc_parts(value)[0]


def _parse_flat_ipc_max_flat(value: object) -> object:
    return _parse_flat_ipc_parts(value)[1]


def _build_flat_ipc_cost_block() -> MatrixCostBlock:
    return MatrixCostBlock(
        cost_name=FLAT_IPC_COLUMN,
        currency_column="Currency Inbound",
        columns=[
            MatrixCostColumn(
                cost_name=FLAT_IPC_COLUMN,
                source_column=FLAT_IPC_COLUMN,
                bracket_label="",
                rate_unit="p/unit",
                parse_value=_parse_flat_ipc_per_unit,
            ),
            MatrixCostColumn(
                cost_name=FLAT_IPC_COLUMN,
                source_column=FLAT_IPC_COLUMN,
                bracket_label="",
                rate_unit="Flat",
                header_label="MAX Flat",
                parse_value=_parse_flat_ipc_max_flat,
            ),
        ],
    )


def _cost_block_has_data(block: MatrixCostBlock, rate_card: pd.DataFrame) -> bool:
    if block.volume_weight_ratio is not None or block.drop_shipment_only:
        return True

    for _, row in rate_card.iterrows():
        for column in block.columns:
            if not _is_empty_value(row[column.source_column]):
                return True
    return False


def _filter_empty_cost_blocks(
    blocks: list[MatrixCostBlock],
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    return [block for block in blocks if _cost_block_has_data(block, rate_card)]


def _column_has_meaningful_cost_data(
    rate_card: pd.DataFrame,
    block: MatrixCostBlock,
    column: MatrixCostColumn,
) -> bool:
    for _, row in rate_card.iterrows():
        value = _value_for_block(row, block, column)
        if _is_empty_value(value) or _is_zero_like(value):
            continue
        return True
    return False


def _drop_empty_or_zero_cost_columns(
    blocks: list[MatrixCostBlock],
    rate_card: pd.DataFrame,
) -> list[MatrixCostBlock]:
    pruned: list[MatrixCostBlock] = []
    for block in blocks:
        if _is_dgr_fee_block(block):
            if _cost_block_has_data(block, rate_card):
                pruned.append(block)
            continue

        kept_columns = [
            column
            for column in block.columns
            if _column_has_meaningful_cost_data(rate_card, block, column)
        ]
        if not kept_columns:
            continue
        pruned.append(
            MatrixCostBlock(
                cost_name=block.cost_name,
                currency_column=block.currency_column,
                columns=kept_columns,
                volume_weight_ratio=block.volume_weight_ratio,
                drop_shipment_only=block.drop_shipment_only,
            )
        )
    return pruned


def _expand_block_columns(block: MatrixCostBlock) -> list[MatrixCostColumn]:
    if _is_dgr_fee_block(block):
        return block.columns

    expanded = [
        MatrixCostColumn(
            cost_name=block.cost_name,
            source_column=block.currency_column,
            bracket_label="Currency",
            rate_unit="Currency",
            is_currency=True,
        )
    ]
    expanded.extend(block.columns)
    return expanded


def _value_for_block(
    row: pd.Series,
    block: MatrixCostBlock,
    column: MatrixCostColumn,
) -> object:
    if not _row_matches_block(row, block):
        return None
    if column.source_column not in row.index:
        return None

    value = row[column.source_column]
    if _is_transport_cost_block(block):
        value = _resolve_transport_column_value(row, column.source_column)
    elif column.parse_value is not None:
        value = column.parse_value(value)

    return _sanitize_cost_value(value, allow_string=column.is_currency)


def _build_matrix_rows(
    rate_card: pd.DataFrame,
    shipment_columns: list[ShipmentColumn],
    cost_blocks: list[MatrixCostBlock],
) -> tuple[list[list[object]], list[MatrixCostColumn]]:
    shipment_headers = [column.header for column in shipment_columns]
    cost_columns = [
        column
        for block in cost_blocks
        for column in _expand_block_columns(block)
    ]

    header_rows: list[list[object]] = []
    for header_index in range(5):
        row_values: list[object] = []
        if header_index == 4:
            row_values.extend(shipment_headers)
        else:
            row_values.extend([""] * len(shipment_headers))

        for block in cost_blocks:
            block_columns = _expand_block_columns(block)
            for column in block_columns:
                if header_index == 0:
                    row_values.append(block.cost_name)
                elif header_index in (1, 2):
                    row_values.append("")
                elif header_index == 3:
                    row_values.append("" if column.is_currency else column.bracket_label)
                elif header_index == 4:
                    if column.is_currency:
                        row_values.append(column.rate_unit)
                    else:
                        row_values.append(column.header_label or column.rate_unit)

        header_rows.append(row_values)

    data_rows: list[list[object]] = []
    for _, row in rate_card.iterrows():
        data_row: list[object] = []
        for column in shipment_columns:
            if column.source_column is None:
                data_row.append(None)
            else:
                data_row.append(
                    _format_shipment_cell_value(
                        column.header,
                        row[column.source_column],
                    )
                )

        for block in cost_blocks:
            for column in _expand_block_columns(block):
                data_row.append(_value_for_block(row, block, column))

        data_rows.append(data_row)

    return header_rows + data_rows, cost_columns


def _merge_cost_name_cells(
    worksheet,
    shipment_column_count: int,
    cost_blocks: list[MatrixCostBlock],
) -> None:
    current_column = shipment_column_count + 1
    for block in cost_blocks:
        block_width = len(_expand_block_columns(block))
        if block_width > 1:
            worksheet.merge_cells(
                start_row=COST_NAME_ROW,
                start_column=current_column,
                end_row=COST_NAME_ROW,
                end_column=current_column + block_width - 1,
            )
        current_column += block_width


def _apply_column_widths(
    worksheet,
    shipment_columns: list[ShipmentColumn],
    cost_columns: list[MatrixCostColumn],
) -> None:
    shipment_column_count = len(shipment_columns)

    for column_index, shipment_column in enumerate(shipment_columns, start=1):
        letter = get_column_letter(column_index)
        header = shipment_column.header

        if header == "ID":
            width = 26
        elif header in DATE_SHIPMENT_COLUMNS:
            width = 12
        elif header in {"Remarks", "additional information", "service code extended"}:
            width = 20
        elif header in BOLD_SHIPMENT_HEADERS:
            width = 16
        else:
            width = 14

        worksheet.column_dimensions[letter].width = width

    for offset, cost_column in enumerate(cost_columns, start=1):
        letter = get_column_letter(shipment_column_count + offset)
        worksheet.column_dimensions[letter].width = 10 if cost_column.is_currency else 11


def _apply_row_heights(worksheet, total_rows: int) -> None:
    for row_index in range(1, DATA_START_ROW):
        worksheet.row_dimensions[row_index].height = HEADER_ROW_HEIGHT

    for row_index in range(DATA_START_ROW, total_rows + 1):
        worksheet.row_dimensions[row_index].height = DATA_ROW_HEIGHT


def _apply_worksheet_formatting(
    worksheet,
    shipment_columns: list[ShipmentColumn],
    cost_columns: list[MatrixCostColumn],
    total_rows: int,
) -> None:
    shipment_column_count = len(shipment_columns)
    total_columns = shipment_column_count + len(cost_columns)
    header_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for row_index in range(1, total_rows + 1):
        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = THIN_BORDER

            if row_index == COST_NAME_ROW and column_index > shipment_column_count:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = header_center
            elif row_index in (APPLY_IF_ROW, RATE_BY_ROW):
                cell.alignment = center
            elif row_index == WEIGHT_BRACKET_ROW and column_index > shipment_column_count:
                cell.fill = SUBHEADER_FILL
                cell.font = SUBHEADER_FONT
                cell.alignment = center
            elif row_index == COLUMN_HEADER_ROW:
                cell.alignment = center
                if column_index <= shipment_column_count and shipment_columns[
                    column_index - 1
                ].bold_header:
                    cell.font = BOLD_FONT
            elif row_index >= DATA_START_ROW and column_index <= shipment_column_count:
                shipment_column = shipment_columns[column_index - 1]
                cell.alignment = left
                if _is_date_shipment_column(shipment_column.header):
                    cell.number_format = DATE_NUMBER_FORMAT
            elif row_index >= DATA_START_ROW and column_index > shipment_column_count:
                cost_column = cost_columns[column_index - shipment_column_count - 1]
                cell.alignment = center
                if not cost_column.is_currency:
                    cell.number_format = PRICE_NUMBER_FORMAT

    _apply_row_heights(worksheet, total_rows)
    _apply_column_widths(worksheet, shipment_columns, cost_columns)


def build_rate_card_matrix(
    rate_card: pd.DataFrame,
    output_path: Path,
    *,
    drop_empty_or_zero_cost_columns: bool = False,
    ignore_volume_weight_ratio: bool = False,
    include_dgr_fee_shipment_column: bool = False,
    sheet_name: str = "Rate Card",
    append_sheet_if_exists: bool = False,
) -> MatrixBuildResult:
    shipment_columns = _build_shipment_columns(rate_card.columns, rate_card)
    if include_dgr_fee_shipment_column:
        shipment_columns = _insert_dgr_fee_shipment_column(shipment_columns, rate_card)
    cost_blocks = _build_cost_blocks(
        rate_card.columns,
        rate_card,
        ignore_volume_weight_ratio=ignore_volume_weight_ratio,
    )
    cost_blocks = _filter_empty_cost_blocks(cost_blocks, rate_card)
    if drop_empty_or_zero_cost_columns:
        cost_blocks = _drop_empty_or_zero_cost_columns(cost_blocks, rate_card)
    matrix_rows, cost_columns = _build_matrix_rows(
        rate_card,
        shipment_columns,
        cost_blocks,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if append_sheet_if_exists and output_path.exists():
        workbook = load_workbook(output_path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

    for row_index, row_values in enumerate(matrix_rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if value == "":
                cell.value = None

    _merge_cost_name_cells(worksheet, len(shipment_columns), cost_blocks)
    _apply_worksheet_formatting(
        worksheet,
        shipment_columns,
        cost_columns,
        len(matrix_rows),
    )

    workbook.save(output_path)

    return MatrixBuildResult(
        matrix_path=output_path,
        row_count=len(rate_card),
        shipment_column_count=len(shipment_columns),
        cost_block_count=len(cost_blocks),
    )


def load_extracted_rate_card(processing_file: Path) -> pd.DataFrame:
    return pd.read_excel(processing_file, sheet_name="Rate Card", header=0)


def _normalize_country_code(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def filter_lanes_ex_to_country(
    rate_card: pd.DataFrame,
    country_code: str,
) -> pd.DataFrame:
    if ORIGIN_COUNTRY_CODE_COLUMN not in rate_card.columns:
        raise ValueError(f"Missing column: {ORIGIN_COUNTRY_CODE_COLUMN}")
    if DESTINATION_COUNTRY_CODE_COLUMN not in rate_card.columns:
        raise ValueError(f"Missing column: {DESTINATION_COUNTRY_CODE_COLUMN}")

    country_code = country_code.strip().upper()
    origin_codes = rate_card[ORIGIN_COUNTRY_CODE_COLUMN].map(_normalize_country_code)
    destination_codes = rate_card[DESTINATION_COUNTRY_CODE_COLUMN].map(
        _normalize_country_code
    )
    keep_mask = (origin_codes == country_code) | (destination_codes == country_code)
    return rate_card[keep_mask].reset_index(drop=True)


def apply_lane_country_filters(
    rate_card: pd.DataFrame,
    *,
    de_only: bool = False,
    latam_only: bool = False,
) -> pd.DataFrame:
    filtered = rate_card
    if de_only:
        filtered = filter_lanes_ex_to_country(filtered, "DE")
    if latam_only:
        latam_codes = set(LATAM_COUNTRY_CODES)
        origin_codes = filtered[ORIGIN_COUNTRY_CODE_COLUMN].map(_normalize_country_code)
        destination_codes = filtered[DESTINATION_COUNTRY_CODE_COLUMN].map(
            _normalize_country_code
        )
        keep_mask = origin_codes.isin(latam_codes) | destination_codes.isin(latam_codes)
        filtered = filtered[keep_mask].reset_index(drop=True)
    return filtered


def prompt_yes_no(question: str) -> bool:
    while True:
        raw = input(f"{question} (yes/no): ").strip().lower()
        if raw in {"yes", "y"}:
            return True
        if raw in {"no", "n"}:
            return False
        print("Please answer yes or no.")


def prompt_lane_country_filters() -> tuple[bool, bool, bool]:
    de_only = prompt_yes_no("Use lanes ex/to DE ONLY?")
    if de_only:
        return True, False, True
    latam_only = prompt_yes_no("Use lanes ex/to LATAM ONLY?")
    drop_empty_or_zero_cost_columns = True
    return False, latam_only, drop_empty_or_zero_cost_columns


def build_matrix_from_rate_card(
    rate_card: pd.DataFrame,
    output_file: Path,
    *,
    de_only: bool = False,
    latam_only: bool = False,
    drop_empty_or_zero_cost_columns: bool = False,
    ignore_volume_weight_ratio: bool = False,
    include_dgr_fee_shipment_column: bool = False,
    sheet_name: str = "Rate Card",
    append_sheet_if_exists: bool = False,
) -> MatrixBuildResult:
    original_row_count = len(rate_card)
    filtered_rate_card = apply_lane_country_filters(
        rate_card,
        de_only=de_only,
        latam_only=latam_only,
    )
    if len(filtered_rate_card) == 0:
        filters = []
        if de_only:
            filters.append("DE")
        if latam_only:
            filters.append(f"LATAM ({', '.join(LATAM_COUNTRY_CODES)})")
        filter_text = " and ".join(filters) if filters else "selected"
        raise ValueError(
            f"No lanes remain after applying {filter_text} country filter(s)."
        )

    result = build_rate_card_matrix(
        filtered_rate_card,
        output_file,
        drop_empty_or_zero_cost_columns=drop_empty_or_zero_cost_columns,
        ignore_volume_weight_ratio=ignore_volume_weight_ratio,
        include_dgr_fee_shipment_column=include_dgr_fee_shipment_column,
        sheet_name=sheet_name,
        append_sheet_if_exists=append_sheet_if_exists,
    )
    if de_only or latam_only:
        print(f"Filtered lanes: {original_row_count} -> {len(filtered_rate_card)} rows")
        if de_only:
            print("  Applied filter: ex/to DE ONLY")
        if latam_only:
            print(
                "  Applied filter: ex/to LATAM ONLY "
                f"({', '.join(LATAM_COUNTRY_CODES)})"
            )
    if drop_empty_or_zero_cost_columns:
        print("  Applied filter: drop empty/zero cost columns")
    return result


def build_matrix_from_processing_file(
    processing_file: Path,
    output_file: Path | None = None,
    *,
    de_only: bool = False,
    latam_only: bool = False,
    drop_empty_or_zero_cost_columns: bool = False,
    ignore_volume_weight_ratio: bool = False,
    include_dgr_fee_shipment_column: bool = False,
) -> MatrixBuildResult:
    rate_card = load_extracted_rate_card(processing_file)
    original_row_count = len(rate_card)
    rate_card = apply_lane_country_filters(
        rate_card,
        de_only=de_only,
        latam_only=latam_only,
    )
    if len(rate_card) == 0:
        filters = []
        if de_only:
            filters.append("DE")
        if latam_only:
            filters.append(f"LATAM ({', '.join(LATAM_COUNTRY_CODES)})")
        filter_text = " and ".join(filters) if filters else "selected"
        raise ValueError(
            f"No lanes remain after applying {filter_text} country filter(s)."
        )
    if output_file is None:
        output_file = OUTPUT_DIR / f"{processing_file.stem.replace('_extracted', '')}_matrix.xlsx"
    result = build_rate_card_matrix(
        rate_card,
        output_file,
        drop_empty_or_zero_cost_columns=drop_empty_or_zero_cost_columns,
        ignore_volume_weight_ratio=ignore_volume_weight_ratio,
        include_dgr_fee_shipment_column=include_dgr_fee_shipment_column,
    )
    if de_only or latam_only:
        print(f"Filtered lanes: {original_row_count} -> {len(rate_card)} rows")
        if de_only:
            print("  Applied filter: ex/to DE ONLY")
        if latam_only:
            print(
                "  Applied filter: ex/to LATAM ONLY "
                f"({', '.join(LATAM_COUNTRY_CODES)})"
            )
    if drop_empty_or_zero_cost_columns:
        print("  Applied filter: drop empty/zero cost columns")
    return result


def list_processing_files() -> list[Path]:
    if not PROCESSING_DIR.exists():
        return []
    return sorted(PROCESSING_DIR.glob("*_extracted.xlsx"))


def run_interactive_matrix_build() -> MatrixBuildResult:
    processing_files = list_processing_files()
    if not processing_files:
        raise FileNotFoundError(f"No extracted files found in {PROCESSING_DIR}")

    print("Extracted files in processing folder:")
    for index, file_path in enumerate(processing_files, start=1):
        print(f"  {index}. {file_path.name}")

    while True:
        raw = input("Enter file number to convert: ").strip()
        if raw.isdigit():
            choice_index = int(raw) - 1
            if 0 <= choice_index < len(processing_files):
                selected_file = processing_files[choice_index]
                break
        print("Invalid choice. Try again.")

    de_only, latam_only, drop_empty_or_zero_cost_columns = prompt_lane_country_filters()
    result = build_matrix_from_processing_file(
        selected_file,
        de_only=de_only,
        latam_only=latam_only,
        drop_empty_or_zero_cost_columns=drop_empty_or_zero_cost_columns,
    )
    print(f"\nSaved matrix rate card to: {result.matrix_path}")
    print(f"Data rows: {result.row_count}")
    print(f"Shipment columns: {result.shipment_column_count}")
    print(f"Cost blocks: {result.cost_block_count}")
    return result


if __name__ == "__main__":
    run_interactive_matrix_build()
